#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""五月账单处理脚本 - 三平台合并"""

import csv
import re
import os
import sys
from datetime import datetime
from collections import defaultdict

# ============ 分类规则（按优先级排序，精确匹配优先） ============
CATEGORY_RULES = [
    # 优先精确匹配的分类
    ('🏠 住宿/房租', {'keywords': ['房租', '水电费'], 'merchants': ['李雄英房东', '爱情公寓']}),
    ('🛵 吃饭-外卖', {'keywords': ['外卖', '闪购'], 'merchants': ['淘宝闪购']}),
    ('🍜 吃饭-正餐/食材', {'keywords': ['餐厅', '烧烤', '超市', '水果', '饭', '粉', '面', '馄饨', '饺子', '米粉', '烧鹅', '煲粉', '炒粉', '糖水', '田螺', '龟苓膏', '便利店', '麦当劳'], 'merchants': ['沃尔玛', '李阿姨十年老店', '昊昊水果店', '姚金烧鹅', '妙凤煲粉', '吉村米粉店', '桂航炒粉', '老东江米粉', '永香兴五香烧汁饭', '老味道米粉铺', '苏姨糖水铺', '北水田螺汤', '老梧州龟苓膏', '美宜佳', '麦当劳', '美团', '共橙']}),
    ('🧋 吃饭-奶茶/零食', {'keywords': ['奶茶', '咖啡', '饮料', '茶饮', '零食'], 'merchants': ['1点点', '蜜雪冰城', '瑞幸', '零食好忙', '赵一鸣']}),
    ('💄 购物-美妆', {'keywords': ['护肤', '面膜', '美妆', '毛戈平', '鱼子酱', '洗发水', '去屑', '控油', '赛逸', 'sheveu'], 'merchants': ['毛戈平', '赛逸', 'sheveu', 'ch**舰', 'sh**店']}),
    ('👔 购物-服饰', {'keywords': ['衣服', '裤子', '鞋', '服饰', '短袖', '内裤', '外套', '羽绒服', 'UTO', 'Airsil', 'NO LOGO'], 'merchants': ['NewDecay', 'UTO悠途', '鸭鸭', 'NO LOGO', 'Airsil', 'ut**店', 'ig**2']}),
    ('📦 购物-抖音', {'keywords': ['抖音电商', '抖音生活'], 'merchants': ['抖音电商商家', '抖音生活服务商家']}),
    ('📦 购物-京东', {'keywords': ['京东'], 'merchants': ['京东', '京邦达']}),
    ('📦 购物-拼多多', {'keywords': ['拼多多'], 'merchants': ['拼多多']}),
    ('📦 购物-淘宝', {'keywords': ['淘宝', '天猫'], 'merchants': ['淘宝', '天猫']}),
    ('💰 支付宝小荷包', {'keywords': ['小荷包', '公款吃喝'], 'merchants': ['支付宝小荷包']}),
    ('🚗 交通', {'keywords': ['充电宝', '车费', '电动车', '格林豪泰', '骑行', '滴滴', '洗车', '松果出行', '一车一人', '南网电动'], 'merchants': ['铁塔能源', '格林豪泰', '滴滴出行', '松果出行', '一车一人', '南网电动', '优洗达']}),
    ('🎮 娱乐', {'keywords': ['游戏', '会员', 'App Store', 'DeepSeek', 'API服务'], 'merchants': ['App Store', '上海稀宇科技', '杭州深度求索', 'DeepSeek']}),
    ('💊 医疗健康', {'keywords': ['药', '医院', '过敏', '保健', '保济丸', '鱼腥草', '感冒', '清热解毒'], 'merchants': ['阿里健康', '药房']}),
    ('📚 学习', {'keywords': ['书', '打印', '照相', '教材', '考研', '网课', '开发板', 'STM32', '辅导讲义', '复习全书'], 'merchants': ['雅萌照相馆', '正点原子', '亿卷']}),
]

INVESTMENT_KEYWORDS = ['基金', '定投', '申购', '理财', '余额宝', '零钱通', '投资理财', '蚂蚁财富', '收益发放']
REFUND_KEYWORDS = ['退款', '退款成功', '已退款']

# 特殊退款（金额+商户关键词）
SPECIAL_REFUNDS = [
    {'counterparty_kw': '****', 'desc_kw': 'ipadpro', 'amount': 4200.00},
]

def is_refund(desc, counterparty, amount, status=''):
    """判断是否是退款"""
    # 检查关键词
    text = f"{desc} {counterparty} {status}"
    for kw in REFUND_KEYWORDS:
        if kw in text:
            return True

    # 特殊退款识别
    for refund in SPECIAL_REFUNDS:
        if refund['counterparty_kw'] in counterparty and refund['desc_kw'] in desc.lower() and refund['amount'] == amount:
            return True

    return False

def classify_transaction(desc, counterparty, amount):
    """分类交易"""
    # 检查是否是投资
    text = f"{desc} {counterparty}"
    for kw in INVESTMENT_KEYWORDS:
        if kw in text:
            return '投资'

    # 分类支出 - 先检查商户精确匹配
    for category, rules in CATEGORY_RULES:
        for merchant in rules['merchants']:
            if merchant in counterparty:
                return category

    # 再检查关键词匹配
    for category, rules in CATEGORY_RULES:
        for kw in rules['keywords']:
            if kw in text:
                return category

    return '📎 其他支出'

def parse_alipay(file_path):
    """解析支付宝CSV"""
    transactions = []

    with open(file_path, 'r', encoding='gbk') as f:
        lines = f.readlines()

    # 找到数据开始行
    start_idx = 0
    for i, line in enumerate(lines):
        if '交易时间' in line and '交易对方' in line:
            start_idx = i + 1
            break

    for line in lines[start_idx:]:
        line = line.strip()
        if not line:
            continue

        parts = line.split(',')
        if len(parts) < 8:
            continue

        try:
            date_str = parts[0].split(' ')[0]
            counterparty = parts[2]
            desc = parts[4]
            direction = parts[5]
            amount_str = parts[6]
            payment = parts[7]
            status = parts[8] if len(parts) > 8 else ''

            # 跳过不计收支和交易关闭的交易
            if direction == '不计收支':
                continue
            if status == '交易关闭':
                continue

            if not amount_str or amount_str == '/':
                continue

            amount = float(amount_str)

            transactions.append({
                'date': date_str,
                'source': '支付宝',
                'counterparty': counterparty,
                'desc': desc,
                'amount': amount,
                'direction': direction,
                'payment': payment,
                'status': status,
            })
        except Exception as e:
            continue

    return transactions

def parse_wechat(file_path):
    """解析微信XLSX"""
    import openpyxl

    transactions = []
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and row[0] == '交易时间':
            header_row = i
            break

    if header_row is None:
        wb.close()
        return transactions

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue

        try:
            if not row or not row[0]:
                continue

            date_val = row[0]
            counterparty = str(row[2]) if row[2] else ''
            desc = str(row[3]) if row[3] else ''
            direction = str(row[4]) if row[4] else ''
            amount = float(row[5]) if row[5] else 0
            payment = str(row[6]) if row[6] else ''
            status = str(row[7]) if row[7] else ''

            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val).split(' ')[0]

            if direction not in ['收入', '支出']:
                continue

            # 跳过已退款的交易
            if '退款' in status:
                continue

            transactions.append({
                'date': date_str,
                'source': '微信',
                'counterparty': counterparty,
                'desc': desc,
                'amount': amount,
                'direction': direction,
                'payment': payment,
                'status': status,
            })
        except Exception as e:
            continue

    wb.close()
    return transactions

def parse_bank_pdf(file_path, password):
    """解析中国银行PDF"""
    try:
        import PyPDF2

        with open(file_path, 'rb') as f:
            pdf_reader = PyPDF2.PdfReader(f, password=password)
            text = ''
            for page in pdf_reader.pages:
                text += page.extract_text()

        transactions = []
        lines = text.split('\n')
        for line in lines:
            match = re.match(r'(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}:\d{2}\s+([-\d,.]+)\s+(.+?)(?:\s+(.+))?$', line.strip())
            if match:
                date_str = match.group(1)
                amount_str = match.group(2).replace(',', '')
                counterparty = match.group(3).strip()
                desc = match.group(4).strip() if match.group(4) else ''

                amount = float(amount_str)
                direction = '支出' if amount < 0 else '收入'
                amount = abs(amount)

                transactions.append({
                    'date': date_str,
                    'source': '银行',
                    'counterparty': counterparty,
                    'desc': desc,
                    'amount': amount,
                    'direction': direction,
                    'payment': '中国银行储蓄卡',
                    'status': '交易成功',
                })

        return transactions
    except Exception as e:
        print(f"解析银行PDF失败: {e}")
        return []

def process_month(month_dir, year, month):
    """处理月度账单"""
    alipay_file = None
    wechat_file = None
    bank_file = None

    for f in os.listdir(month_dir):
        if f.endswith('.csv') and '支付宝' in f:
            alipay_file = os.path.join(month_dir, f)
        elif f.endswith('.xlsx') and '微信' in f:
            wechat_file = os.path.join(month_dir, f)
        elif f.endswith('.pdf') and '银行' in f:
            bank_file = os.path.join(month_dir, f)

    print(f"支付宝文件: {alipay_file}")
    print(f"微信文件: {wechat_file}")
    print(f"银行文件: {bank_file}")

    all_transactions = []

    if alipay_file:
        print("\n解析支付宝账单...")
        alipay_txns = parse_alipay(alipay_file)
        print(f"  解析到 {len(alipay_txns)} 条交易")
        all_transactions.extend(alipay_txns)

    if wechat_file:
        print("\n解析微信账单...")
        wechat_txns = parse_wechat(wechat_file)
        print(f"  解析到 {len(wechat_txns)} 条交易")
        all_transactions.extend(wechat_txns)

    if bank_file:
        print("\n解析银行账单...")
        password_match = re.search(r'密码(\d+)', os.path.basename(bank_file))
        password = password_match.group(1) if password_match else None

        if password:
            bank_txns = parse_bank_pdf(bank_file, password)
            print(f"  解析到 {len(bank_txns)} 条交易")
            all_transactions.extend(bank_txns)
        else:
            print("  未找到密码，跳过银行账单")

    # 去重和清洗
    print("\n清洗数据...")
    cleaned = []
    seen = set()

    for txn in all_transactions:
        # 剔除金额为0的交易
        if txn['amount'] == 0:
            continue

        # 去重（日期+金额+交易对方+来源）
        key = (txn['date'], txn['amount'], txn['counterparty'], txn['source'])
        if key in seen:
            continue
        seen.add(key)

        cleaned.append(txn)

    print(f"  去重后剩余 {len(cleaned)} 条交易（剔除 {len(all_transactions) - len(cleaned)} 条）")

    return cleaned

def generate_report(transactions, year, month):
    """生成报告"""
    income = {'生活费/转账收入': [], '其他收入': []}
    expense = defaultdict(list)
    investment = []
    refund = []

    platform_stats = defaultdict(lambda: {'income': 0, 'expense': 0, 'count': 0})

    for txn in transactions:
        date = txn['date']
        amount = txn['amount']
        source = txn['source']
        direction = txn['direction']
        desc = txn['desc']
        counterparty = txn['counterparty']
        payment = txn['payment']
        status = txn.get('status', '')

        # 检查是否是退款
        if is_refund(desc, counterparty, amount, status):
            refund.append(txn)
            continue

        # 分类交易
        category = classify_transaction(desc, counterparty, amount)

        if category == '投资':
            investment.append(txn)
            continue

        if direction == '收入':
            if '红包' in counterparty or '拼多多' in counterparty:
                income['其他收入'].append(txn)
            else:
                income['生活费/转账收入'].append(txn)
            platform_stats[source]['income'] += amount
        elif direction == '支出':
            txn['category'] = category
            expense[category].append(txn)
            platform_stats[source]['expense'] += amount

        platform_stats[source]['count'] += 1

    total_income = sum(t['amount'] for cat in income.values() for t in cat)
    total_expense = sum(t['amount'] for cat in expense.values() for t in cat)
    net_balance = total_income - total_expense

    report = f"""# {month}月账单逐笔明细

> 统计时间：{year}年{month}月1日 - {month}月31日
> 说明：已退款项目已全部剔除，投资类交易不计入
> 数据来源：支付宝、微信、中国银行

---

## 一、收入明细

### 生活费/转账收入（小计 ¥{sum(t['amount'] for t in income['生活费/转账收入']):.2f}）

| 日期 | 金额 | 来源 | 交易对方 | 备注 |
| ---- | ---: | :--: | -------- | ---- |
"""

    for t in sorted(income['生活费/转账收入'], key=lambda x: x['date']):
        report += f"| {t['date']} | ¥{t['amount']:.2f} | {t['source']} | {t['counterparty']} | {t.get('desc', '')} |\n"

    report += f"""
### 其他收入（小计 ¥{sum(t['amount'] for t in income['其他收入']):.2f}）

| 日期 | 金额 | 来源 | 交易对方 | 备注 |
| ---- | ---: | :--: | -------- | ---- |
"""

    for t in sorted(income['其他收入'], key=lambda x: x['date']):
        report += f"| {t['date']} | ¥{t['amount']:.2f} | {t['source']} | {t['counterparty']} | {t.get('desc', '')} |\n"

    report += f"\n**收入合计：¥{total_income:.2f}**\n\n---\n\n## 二、支出明细\n"

    category_order = [
        '🏠 住宿/房租', '🛵 吃饭-外卖', '🍜 吃饭-正餐/食材', '🧋 吃饭-奶茶/零食',
        '💄 购物-美妆', '👔 购物-服饰', '📦 购物-抖音', '📦 购物-京东', '📦 购物-拼多多', '📦 购物-淘宝',
        '💰 支付宝小荷包', '🚗 交通', '🎮 娱乐', '💊 医疗健康', '📚 学习', '📎 其他支出'
    ]

    for cat in category_order:
        txns = expense.get(cat, [])
        if not txns:
            continue

        total = sum(t['amount'] for t in txns)
        report += f"\n### {cat}（小计 ¥{total:.2f}）\n\n"
        report += "| 日期 | 金额 | 来源 | 交易对方 | 商品/说明 | 支付方式 |\n"
        report += "| ---- | ---: | :--: | -------- | --------- | -------- |\n"

        for t in sorted(txns, key=lambda x: x['date']):
            report += f"| {t['date']} | ¥{t['amount']:.2f} | {t['source']} | {t['counterparty']} | {t['desc']} | {t['payment']} |\n"

    report += f"\n**支出合计：¥{total_expense:.2f}**\n"

    report += f"""
---

## 三、汇总

| 项目 | 金额 |
| ---- | ---: |
| 总收入 | ¥{total_income:.2f} |
| 总支出 | ¥{total_expense:.2f} |
| **净结余** | **¥{net_balance:.2f}** |

---

## 四、支出分类汇总

| 分类 | 金额 | 占比 | 笔数 |
| ---- | ---: | ---: | ---: |
"""

    for cat in category_order:
        txns = expense.get(cat, [])
        if txns:
            total = sum(t['amount'] for t in txns)
            percent = (total / total_expense * 100) if total_expense > 0 else 0
            report += f"| {cat} | ¥{total:.2f} | {percent:.1f}% | {len(txns)} |\n"

    report += f"""
---

## 五、各平台收支统计

| 平台 | 收入 | 支出 | 笔数 |
| ---- | ---: | ---: | ---: |
"""

    for platform, stats in sorted(platform_stats.items()):
        report += f"| {platform} | ¥{stats['income']:.2f} | ¥{stats['expense']:.2f} | {stats['count']} |\n"

    if investment:
        report += f"""
---

## 六、投资明细（不计入支出）

| 日期 | 金额 | 来源 | 投资类型 | 说明 |
| ---- | ---: | :--: | -------- | ---- |
"""
        for t in sorted(investment, key=lambda x: x['date']):
            report += f"| {t['date']} | ¥{t['amount']:.2f} | {t['source']} | 投资理财 | {t['desc']} |\n"

        total_investment = sum(t['amount'] for t in investment)
        report += f"\n**投资合计：¥{total_investment:.2f}**\n"

    if refund:
        report += f"""
---

## 七、退款明细（已剔除）

| 日期 | 金额 | 来源 | 交易对方 | 说明 |
| ---- | ---: | :--: | -------- | ---- |
"""
        for t in sorted(refund, key=lambda x: x['date']):
            report += f"| {t['date']} | ¥{t['amount']:.2f} | {t['source']} | {t['counterparty']} | {t['desc']} |\n"

        total_refund = sum(t['amount'] for t in refund)
        report += f"\n**退款合计：¥{total_refund:.2f}**\n"

    report += f"\n---\n\n*报告生成时间：{datetime.now().strftime('%Y年%m月%d日')}*\n"

    return report

if __name__ == '__main__':
    if len(sys.argv) >= 4:
        month_dir = sys.argv[1]
        year = int(sys.argv[2])
        month = int(sys.argv[3])
    else:
        month_dir = r'G:\1_Knowledge\Obsidian\记账\五月'
        year = 2026
        month = 5

    print(f"开始处理{year}年{month}月账单（三平台合并）...")
    transactions = process_month(month_dir, year, month)

    if transactions:
        print(f"\n共解析 {len(transactions)} 条交易")

        report = generate_report(transactions, year, month)

        output_file = f'{month_dir}/{month}月账单逐笔明细.md'
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(report)

        print(f"\n报告已保存到: {output_file}")
    else:
        print("未找到交易记录")
